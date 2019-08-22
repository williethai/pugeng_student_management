from django.shortcuts import render
from django.views.generic import TemplateView
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from rest_framework import generics
from rest_framework.generics import GenericAPIView
from rest_framework.mixins import UpdateModelMixin
from rest_framework.viewsets import ModelViewSet
from rest_framework import serializers

from datetime import datetime, date, timedelta
from dateutil.relativedelta import *

import json
import csv
from openpyxl import Workbook

from .models import *

def export_student_csv(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    students = Student.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-students.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = '學員'

    # Define the titles for columns
    columns = []
    columns.append('姓名')
    columns.append('法名')
    columns.append('生日')
    columns.append('電話')

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for student_i in students:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            student_i.name,
            student_i.clerical_name,
            student_i.date_of_birth,
            student_i.phone_num,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response